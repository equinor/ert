"""Testing excel_to_dict"""

import numpy as np
import openpyxl
import pandas as pd
import pytest

from fmudesign import excel_to_dict, inputdict_to_yaml
from fmudesign._excel_to_dict import (
    _assert_no_merged_cells,
    _has_value,
)

MOCK_GENERAL_INPUT = pd.DataFrame(
    data=[
        ["designtype", "onebyone"],
        ["repeats", "10"],
        ["rms_seeds", "default"],
        ["background", "None"],
        ["distribution_seed", 42],
    ]
)

MOCK_DESIGNINPUT = pd.DataFrame(
    data=[["sensname", "numreal", "type", "param_name"], ["rms_seed", "", "seed"]]
)


def _write_config_workbook(
    path,
    *,
    general_input=MOCK_GENERAL_INPUT,
    design_input=MOCK_DESIGNINPUT,
    defaultvalues=None,
    background=None,
    general_sheet="general_input",
    design_sheet="designinput",
    default_sheet="defaultvalues",
    background_sheet="backgroundsheet",
):
    if defaultvalues is None:
        defaultvalues = pd.DataFrame()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        general_input.to_excel(
            writer, sheet_name=general_sheet, index=False, header=None
        )
        design_input.to_excel(writer, sheet_name=design_sheet, index=False, header=None)
        defaultvalues.to_excel(
            writer, sheet_name=default_sheet, index=False, header=None
        )
        if background is not None:
            background.to_excel(
                writer, sheet_name=background_sheet, index=False, header=None
            )
    return path


def test_that_excel_to_dict_parses_workbook_into_configuration_dictionary(tmp_path):
    input_path = _write_config_workbook(tmp_path / "designinput.xlsx")
    dict_design = excel_to_dict(input_path)

    assert isinstance(dict_design, dict)
    assert dict_design["designtype"] == "onebyone"
    assert dict_design["distribution_seed"] == 42
    assert dict_design["defaultvalues"] == {}
    assert isinstance(dict_design["sensitivities"], dict)

    sens = dict_design["sensitivities"]
    assert sens["rms_seed"]["seedname"] == "RMS_SEED"
    assert sens["rms_seed"]["senstype"] == "seed"

    alternate_path = _write_config_workbook(
        tmp_path / "designinput2.xlsx",
        general_sheet="Generalinput",
        design_sheet="Design_input",
        default_sheet="DefaultValues",
    )
    dict_design = excel_to_dict(alternate_path)
    assert isinstance(dict_design, dict)
    assert dict_design["sensitivities"]["rms_seed"]["senstype"] == "seed"

    yaml_path = tmp_path / "dictdesign.yaml"
    inputdict_to_yaml(dict_design, yaml_path)
    assert "RMS_SEED" in yaml_path.read_text(encoding="utf-8")


def test_that_duplicate_sensitivity_names_raise_value_error(tmp_path):
    mock_erroneous_designinput = pd.DataFrame(
        data=[
            ["sensname", "numreal", "type", "param_name"],
            ["rms_seed", "", "seed"],
            ["rms_seed", "", "seed"],
            [np.nan, "", "seed"],  # NaN sensname - should be ignored
            ["", "", "seed"],  # Empty string - should be ignored
            ["valid_name", "", "seed"],  # Valid unique name
        ]
    )
    input_path = _write_config_workbook(
        tmp_path / "designinput.xlsx", design_input=mock_erroneous_designinput
    )

    with pytest.raises(
        ValueError, match="Two sensitivities cannot share the same sensname"
    ):
        excel_to_dict(input_path)


def test_that_excel_to_dict_strips_sensitivity_and_parameter_name_whitespace(
    tmp_path,
):
    """Spaces before and after parameter names are probably
    invisible user errors in Excel sheets. Remove them.
    """
    # pylint: disable=abstract-class-instantiated
    mock_spacious_designinput = pd.DataFrame(
        data=[
            ["sensname", "numreal", "type", "param_name"],
            ["rms_seed   ", "", "seed"],
        ]
    )
    defaultvalues_spacious = pd.DataFrame(
        data=[
            ["parametername", "value"],
            ["  spacious_multiplier", 1.2],
            ["spacious2  ", 3.3],
        ]
    )
    input_path = _write_config_workbook(
        tmp_path / "designinput.xlsx",
        design_input=mock_spacious_designinput,
        defaultvalues=defaultvalues_spacious,
    )

    dict_design = excel_to_dict(input_path)
    assert next(iter(dict_design["sensitivities"].keys())) == "rms_seed"
    def_params = list(dict_design["defaultvalues"].keys())
    assert [par.strip() for par in def_params] == def_params


def test_that_mixed_sensitivity_types_raise_value_error(tmp_path):
    mock_erroneous_designinput = pd.DataFrame(
        data=[
            ["sensname", "numreal", "type", "param_name"],
            ["rms_seed", "", "seed"],
            ["", "", "dist"],
        ]
    )
    input_path = _write_config_workbook(
        tmp_path / "designinput.xlsx", design_input=mock_erroneous_designinput
    )

    with pytest.raises(ValueError, match="contains more than one sensitivity type"):
        excel_to_dict(input_path)


@pytest.mark.parametrize(
    ("value", "expected"), [(1, True), (np.nan, False), (None, True)]
)
def test_that_has_value_treats_only_nan_as_missing(value, expected):
    assert _has_value(value) is expected


def test_that_excel_to_dict_parses_background_sheet(tmp_path):
    general_input = pd.DataFrame(
        data=[
            ["designtype", "onebyone"],
            ["repeats", 3],
            ["rms_seeds", "default"],
            ["background", "backgroundsheet"],
            ["distribution_seed", 42],
        ]
    )
    defaultvalues = pd.DataFrame(
        data=[["param_name", "default_value"], ["extraseed", "0"]]
    )
    background = pd.DataFrame(
        data=[
            ["param_name", "dist_name", "dist_param1"],
            ["extraseed", "scenario", "30,40,50"],
        ]
    )

    input_path = _write_config_workbook(
        tmp_path / "designinput.xlsx",
        general_input=general_input,
        defaultvalues=defaultvalues,
        background=background,
        design_sheet="design_input",
    )

    dict_design = excel_to_dict(input_path)

    # Assert it has been interpreted correctly from input files:
    assert dict_design["background"]["parameters"]["extraseed"] == [
        "scenario",
        ["30,40,50"],
        None,
    ]
    assert dict_design["repeats"] == 3
    assert dict_design["defaultvalues"]["extraseed"] == 0


def _write_background_workbook(path, background, corr_matrix, background_name):
    general_input = pd.DataFrame(
        data=[
            ["designtype", "onebyone"],
            ["repeats", 3],
            ["rms_seeds", "default"],
            ["background", background_name],
            ["distribution_seed", 42],
        ]
    )
    defaultvalues = pd.DataFrame(
        columns=["param_name", "default_value"], data=[["PARAM_A", 0], ["PARAM_B", 0]]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        general_input.to_excel(
            writer, sheet_name="general_input", index=False, header=None
        )
        MOCK_DESIGNINPUT.to_excel(
            writer, sheet_name="designinput", index=False, header=None
        )
        defaultvalues.to_excel(writer, sheet_name="defaultvalues", index=False)
        background.to_excel(
            writer, sheet_name="backgroundsheet", index=False, header=None
        )
        if corr_matrix is not None:
            corr_matrix.to_excel(writer, sheet_name="bgcorr")
    return path


BACKGROUND_WITH_CORR = pd.DataFrame(
    data=[
        ["param_name", "dist_name", "dist_param1", "dist_param2", "corr_sheet"],
        ["PARAM_A", "uniform", 0, 1, "bgcorr"],
        ["PARAM_B", "uniform", 0, 1, "bgcorr"],
    ]
)


@pytest.mark.parametrize(
    ("index", "columns", "error"),
    [
        (
            ["PARAM_A", "PARAM_B"],
            ["PARAM_A", "PARAM_TYPO"],
            "Mismatch between column and index in correlation",
        ),
        (
            ["PARAM_A", "PARAM_TYPO"],
            ["PARAM_A", "PARAM_TYPO"],
            "Mismatch between parameters",
        ),
    ],
)
def test_that_invalid_background_correlation_sheet_raises_value_error(
    tmp_path, index, columns, error
):
    corr_matrix = pd.DataFrame(
        [[1.0, np.nan], [0.5, 1.0]],
        index=index,
        columns=columns,
    )
    input_path = _write_background_workbook(
        tmp_path / "designinput.xlsx",
        BACKGROUND_WITH_CORR,
        corr_matrix,
        "backgroundsheet",
    )

    with pytest.raises(ValueError, match=error):
        excel_to_dict(input_path)


def test_that_missing_background_sheet_error_lists_available_sheets(tmp_path):
    input_path = _write_background_workbook(
        tmp_path / "designinput.xlsx", BACKGROUND_WITH_CORR, None, "typo_sheet"
    )

    with pytest.raises(ValueError, match="Sheets in workbook") as exc_info:
        excel_to_dict(input_path)

    message = str(exc_info.value)
    assert "typo_sheet" in message
    assert "backgroundsheet" in message
    assert "Use 'None' as background" in message


@pytest.mark.parametrize(
    "background_name", ["Backgroundsheet", "background_sheet", " backgroundsheet "]
)
def test_that_background_sheet_matching_ignores_case_underscores_and_whitespace(
    tmp_path, background_name
):
    corr_matrix = pd.DataFrame(
        [[1.0, np.nan], [0.5, 1.0]],
        index=["PARAM_A", "PARAM_B"],
        columns=["PARAM_A", "PARAM_B"],
    )
    input_path = _write_background_workbook(
        tmp_path / "designinput.xlsx",
        BACKGROUND_WITH_CORR,
        corr_matrix,
        background_name,
    )

    background = excel_to_dict(input_path)["background"]
    assert list(background["parameters"]) == ["PARAM_A", "PARAM_B"]
    assert background["correlations"]["sheetnames"] == ["bgcorr"]


def test_that_missing_background_csv_file_raises_value_error(use_tmpdir):
    input_path = _write_background_workbook(
        "designinput.xlsx",
        BACKGROUND_WITH_CORR,
        None,
        "missing_background.csv",
    )

    with pytest.raises(
        ValueError,
        match=r"Sheet 'missing_background.csv' with background parameters, "
        "specified in the general input sheet, "
        "was not found in 'designinput.xlsx'.",
    ):
        excel_to_dict(input_path)


@pytest.mark.parametrize("background_name", ["None", "none", np.nan])
def test_that_none_like_background_names_disable_background(tmp_path, background_name):
    input_path = _write_background_workbook(
        tmp_path / "designinput.xlsx", BACKGROUND_WITH_CORR, None, background_name
    )

    assert excel_to_dict(input_path)["background"] is None


def test_that_assert_no_merged_cells_rejects_merged_cells(tmp_path):
    input_path = tmp_path / "test_file.xlsx"
    test_data = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    test_data.to_excel(input_path, sheet_name="sheet1", index=False)

    workbook = openpyxl.load_workbook(input_path)
    workbook["sheet1"].merge_cells("A1:B1")
    workbook.save(input_path)
    workbook.close()

    with pytest.raises(ValueError, match="Merged cells"):
        _assert_no_merged_cells(input_path)


def test_that_excel_to_dict_preserves_seed_strategy(tmp_path):
    general = pd.DataFrame(
        data=[
            ["designtype", "onebyone"],
            ["repeats", "10"],
            ["rms_seeds", "default"],
            ["background", "None"],
            ["distribution_seed", 42],
            ["seed_strategy", "independent"],
        ]
    )
    designinput = pd.DataFrame(
        data=[["sensname", "numreal", "type", "param_name"], ["rms_seed", "", "seed"]]
    )
    input_path = _write_config_workbook(
        tmp_path / "designinput.xlsx",
        general_input=general,
        design_input=designinput,
    )
    dict_design = excel_to_dict(input_path)
    assert dict_design["seed_strategy"] == "independent"
