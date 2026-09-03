"""This module contains functions for reading Excel config files.
These are converted to a dict-of-dicts representation, then they are used
by the DesignMatrix class to generate design matrices.
"""

from collections import Counter
from collections.abc import Hashable, Sequence
from pathlib import Path
from typing import Any, Literal, cast

import openpyxl
import pandas as pd
import yaml

from .general_input import GeneralInput
from .read_background import read_background
from .read_correlations import parse_sensitivity_correlations
from .utils import (
    _has_value,
    find_sheet,
    resolve_path,
    seeds_from_extern,
)


def excel_to_dict(
    input_filename: str,
    *,
    gen_input_sheet: str = "general_input",
    design_input_sheet: str = "designinput",
    default_val_sheet: str = "defaultvalues",
) -> dict[str, Any]:
    """Read excel file with input to design setup

    Args:
        input_filename (str): Name of excel input file
        gen_input_sheet (str): Sheet name for general input
        design_input_sheet (str): Sheet name for design input
        default_val_sheet (str): Sheet name for default input

    Returns:
        dict on format for DesignMatrix.generate
    """
    # To be backwards compatible, we do not change the input arg names
    general_input_sheet = gen_input_sheet
    default_values_sheet = default_val_sheet

    # Find sheets
    _assert_no_merged_cells(input_filename)
    xlsx = openpyxl.load_workbook(input_filename, read_only=True, keep_links=False)
    general_input_sheet = find_sheet(general_input_sheet, names=xlsx.sheetnames)
    design_input_sheet = find_sheet(design_input_sheet, names=xlsx.sheetnames)
    default_values_sheet = find_sheet(default_values_sheet, names=xlsx.sheetnames)

    general_input = GeneralInput.from_xlsx(input_filename, general_input_sheet)

    return _excel_to_dict_onebyone(
        input_filename=input_filename,
        general_input=general_input,
        design_input_sheet=design_input_sheet,
        default_values_sheet=default_values_sheet,
    )


def inputdict_to_yaml(inputdict: dict[str, Any], filename: str) -> None:
    """Write inputdict to yaml format

    Args:
        inputdict (dict)
        filename (str): name of output file
    """
    with Path(filename).open("w", encoding="utf-8") as stream:
        yaml.dump(inputdict, stream)


def _check_designinput(dsgn_input: pd.DataFrame) -> None:
    """Checks for valid input in designinput sheet"""
    # Filter out rows where sensname has no value
    valid_sensnames = dsgn_input["sensname"].dropna()
    duplicated_mask = valid_sensnames.duplicated()

    if duplicated_mask.any():
        # Find the first duplicate to include in error message
        duplicate_name = valid_sensnames[duplicated_mask].iloc[0]
        raise ValueError(
            f"sensname '{duplicate_name}' was found on more than one row in "
            "designinput sheet. Two sensitivities cannot share the same sensname. "
            "Please correct this and rerun"
        )

    # Check for duplicate parameter names within each sensname
    for sensname, df_sensname in dsgn_input.ffill().groupby("sensname"):
        param_names = list(df_sensname["param_name"])
        try:
            _raise_if_duplicates(param_names)
        except ValueError as e:
            raise ValueError(f"Duplicate param names in {sensname}\n{e}") from e


def _check_for_mixed_sensitivities(sens_name: str, sens_group: pd.DataFrame) -> None:
    """Checks for valid input in designinput sheet. A sensitivity cannot contain
    two different sensitivity types
    """

    unique_types = sens_group["type"].dropna().unique()
    if len(unique_types) > 1:
        raise ValueError(
            f"The sensitivity with sensname '{sens_name}' in designinput sheet "
            "contains more than one sensitivity type. For each sensname all parameters "
            "must be specified using the same type (seed, scenario, dist, ref, "
            "background, extern)"
        )


def _excel_to_dict_onebyone(
    input_filename: str,
    *,
    general_input: GeneralInput,
    design_input_sheet: str,
    default_values_sheet: str,
) -> dict[str, Any]:
    """Reads configuration from Excel file for a onebyone design matrix.

    Args:
        input_filename (str): Name of excel workbook
        general_input (GeneralInput): Validated general input
        design_input_sheet (str): name of design input sheet
        default_values_sheet (str): name of default value sheet

    Returns:
        dict on format for DesignMatrix.generate
    """

    if isinstance(seeds := general_input.rms_seeds, Path):
        rms_seeds: Literal["default"] | list[int] | None = seeds_from_extern(seeds)
    else:
        rms_seeds = seeds

    if isinstance(bgr := general_input.background, Path):
        background = {"extern": str(bgr)}
    elif isinstance(bgr, str):
        background = read_background(input_filename, bgr)
    else:
        background = None

    output: dict[str, Any] = {
        "input_file": input_filename,
        "designtype": general_input.designtype,
        "repeats": general_input.repeats,
        "distribution_seed": general_input.distribution_seed,
        "background": background,
        "seeds": rms_seeds,
        "correlation_iterations": general_input.correlation_iterations,
        "seed_strategy": general_input.seed_strategy,
        "defaultvalues": _read_defaultvalues(input_filename, default_values_sheet),
        "sensitivities": {},
    }  # This is the config that we read and return

    designinput = (
        pd.read_excel(input_filename, design_input_sheet, engine="openpyxl")
        .dropna(axis=0, how="all")
        .loc[:, lambda df: ~df.columns.astype(str).str.contains("^Unnamed")]
    )

    # Strip strings in column 'sensname' while preserving NaN values
    designinput = designinput.assign(sensname=lambda df: df["sensname"].str.strip())

    _check_designinput(designinput)

    designinput["sensname"] = designinput["sensname"].ffill()

    if "decimals" in designinput:
        # Convert to numeric, then filter for integers
        numeric_decimals = pd.to_numeric(designinput["decimals"], errors="coerce")
        mask = numeric_decimals.notna() & (numeric_decimals % 1 == 0)

        valid_decimals = designinput[mask]
        output["decimals"] = {
            row.param_name: int(cast("float", row.decimals))
            for row in valid_decimals.itertuples()
        }

    grouped = designinput.groupby("sensname", sort=False)

    # Read each sensitivity
    for sensname, group in grouped:
        _check_for_mixed_sensitivities(
            str(sensname),
            group,
        )

        sensdict: dict[str, Any] = {}

        sens_type = group["type"].iloc[0]
        if sens_type in {"ref", "background"}:
            sensdict["senstype"] = sens_type

        elif sens_type == "seed":
            sensdict["seedname"] = "RMS_SEED"
            sensdict["senstype"] = sens_type
            if _has_value(group["param_name"].iloc[0]):
                sensdict["parameters"] = _read_constants(group)
            else:
                sensdict["parameters"] = None

        elif sens_type == "scenario":
            sensdict = _read_scenario_sensitivity(group)
            sensdict["senstype"] = sens_type

        elif sens_type == "dist":
            sensdict["senstype"] = sens_type
            sensdict["parameters"] = _read_dist_sensitivity(group)
            sensdict["correlations"] = None
            if "corr_sheet" in group:
                sensdict["correlations"] = parse_sensitivity_correlations(
                    group, input_filename
                )

        elif sens_type == "extern":
            sensdict["extern_file"] = resolve_path(
                str(group["extern_file"].iloc[0]), base_file=input_filename
            )
            sensdict["senstype"] = sens_type
            sensdict["parameters"] = list(group["param_name"])

        else:
            raise ValueError(
                f"Sensitivity {sensname} does not have a valid sensitivity type"
            )

        if "numreal" in group and _has_value(group["numreal"].iloc[0]):
            # Using default number of realisations:
            # 'repeats' from general_input sheet
            sensdict["numreal"] = int(group["numreal"].iloc[0])

        # If this sensitivity has dependencies, then get them from sheet
        sensdict["dependencies"] = {}
        if "dependencies" in group:
            # Get all dependencies in this sensitivity
            valid_deps = group[group["dependencies"].notna()]
            dependencies_dict = {}

            # For each dependency, get the mapping
            for row in valid_deps.itertuples():
                dependencies_dict[row.param_name] = _read_dependencies(
                    filename=input_filename,
                    sheetname=str(row.dependencies),
                    from_parameter=str(row.param_name),
                )
            sensdict["dependencies"] = dependencies_dict

        # Add this sensitivity to the sensitivities
        output["sensitivities"][str(sensname)] = sensdict

    return output


def _read_defaultvalues(filename: str, sheetname: str) -> dict[str, Any]:
    """Reads defaultvalues, also used as values for
    reference/base case

    Args:
        filename (str): Name of excel file
        sheetname (string): name of defaultsheet

    Returns:
        dict with defaultvalues (parameter, value)
    """
    default_df = (
        pd.read_excel(filename, sheetname, header=0, index_col=0, engine="openpyxl")
        .dropna(axis=0, how="all")
        # Drop all unnamed columns from the df
        .loc[:, lambda df: ~df.columns.astype(str).str.contains("^Unnamed")]
    )

    if default_df.empty:
        return {}

    # Strip leading/trailing spaces from parameter names such that
    # for example "  PARAM" and "PARAM" are treated as duplicates.
    default_df.index = default_df.index.str.strip()

    # Check for duplicates and raise error if found
    duplicates = default_df.index.duplicated(keep=False)
    if duplicates.any():
        duplicate_names = default_df.index[duplicates].unique()
        raise ValueError(
            f"Duplicate parameter names found in sheet '{sheetname}': "
            f"{', '.join(duplicate_names)}. All parameter names must be unique."
        )

    return {str(k): v for k, v in default_df.iloc[:, 0].to_dict().items()}


def _read_dependencies(
    *, filename: str, sheetname: str, from_parameter: str
) -> dict[str, Any]:
    """Reads parameters that are set from other parameters

    Args:
        filename(str): name of excel file
        sheetname (string): name of dependency sheet
        from_parameter (string): parameter name to map from

    Returns:
        dict with design parameter, dependent parameters
        and values
    """
    depend_dict: dict[str, Any] = {}
    depend_df = (
        pd.read_excel(filename, sheetname, dtype=str, na_values="", engine="openpyxl")
        .dropna(axis=0, how="all")
        .loc[:, lambda df: ~df.columns.astype(str).str.contains("^Unnamed")]
    )

    if from_parameter in depend_df:
        depend_dict["from_values"] = depend_df[from_parameter].tolist()
        depend_dict["to_params"] = {}
        for key in depend_df:
            if key != from_parameter:
                depend_dict["to_params"][key] = depend_df[key].tolist()
    else:
        raise ValueError(
            f"Parameter {from_parameter} specified to have derived parameters, "
            f"but the sheet specifying the dependencies {sheetname} does "
            "not contain the input parameter. "
        )
    return depend_dict


def _read_scenario_sensitivity(sensgroup: pd.DataFrame) -> dict[str, Any]:
    """Reads parameters and values
    for scenario sensitivities
    """
    sdict: dict[str, Any] = {}
    sdict["cases"] = {}
    casedict1: dict[str, Any] = {}
    casedict2: dict[str, Any] = {}

    if not _has_value(sensgroup["senscase1"].iloc[0]):
        raise ValueError(
            "Sensitivity {} has been input "
            "as a scenario sensitivity, but "
            "without a name in senscase1 column.".format(sensgroup["sensname"].iloc[0])
        )

    for row in sensgroup.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                f"Scenario sensitivity {row.sensname} specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.value1):
            raise ValueError(
                f"Parameter {row.param_name} has been input "
                'as type "scenario" but with empty '
                "value in value1 column "
            )
        casedict1[str(row.param_name)] = row.value1

    if _has_value(sensgroup["senscase2"].iloc[0]):
        for row in sensgroup.itertuples():
            if not _has_value(row.value2):
                raise ValueError(
                    "Sensitivity {} has been input "
                    "with a name in senscase2 column "
                    "but without a value for parameter {} "
                    "in value2 column.".format(
                        sensgroup["sensname"].iloc[0], row.param_name
                    )
                )
            casedict2[str(row.param_name)] = row.value2
        sdict["cases"][str(sensgroup["senscase1"].iloc[0])] = casedict1
        sdict["cases"][str(sensgroup["senscase2"].iloc[0])] = casedict2
    else:
        for row in sensgroup.itertuples():
            if _has_value(row.value2):
                raise ValueError(
                    "Sensitivity {} has been input "
                    "with a value for parameter {} "
                    "in value2 column "
                    "but without a name for the scenario "
                    "in senscase2 column.".format(
                        sensgroup["sensname"].iloc[0], row.param_name
                    )
                )
        sdict["cases"][str(sensgroup["senscase1"].iloc[0])] = casedict1
    return sdict


def _read_constants(sensgroup: pd.DataFrame) -> dict[str, Any]:
    """Reads constants to be used together with
    seed sensitivity
    """
    if "dist_param1" not in sensgroup.columns.to_numpy():
        sensgroup["dist_param1"] = float("NaN")
    paramdict: dict[str, Any] = {}
    for row in sensgroup.itertuples():
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter name {row.param_name} has been input "
                'in a sensitivity of type "seed". \n'
                f"If {row.param_name} was meant to be the name of "
                "the seed parameter, this is "
                "unfortunately not allowed. "
                "The seed parameter name is standardised "
                "to RMS_SEED and should not be specified.\n "
                "If you instead meant to specify a constant "
                "value for another parameter in the seed "
                'sensitivity, please remember "const" in '
                'dist_name and a value in "dist_param1". '
            )
        distparams = row.dist_param1
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams]
    return paramdict


def _read_dist_sensitivity(sensgroup: pd.DataFrame) -> dict[str, Any]:
    """Reads parameters and distributions
    for monte carlo sensitivities
    """
    for col_name in ("dist_param1", "dist_param2", "dist_param3", "dist_param4"):
        if col_name not in sensgroup:
            sensgroup[col_name] = float("NaN")
    paramdict: dict[str, Any] = {}
    for row in sensgroup.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                f"Dist sensitivity {row.sensname} specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter {row.param_name} has been input "
                'as type "dist" but with empty '
                "first distribution parameter "
            )
        if not _has_value(row.dist_param2) and _has_value(row.dist_param3):
            raise ValueError(
                f"Parameter {row.param_name} has been input with "
                'value for "dist_param3" while '
                '"dist_param2" is empty. This is not '
                "allowed"
            )
        if not _has_value(row.dist_param3) and _has_value(row.dist_param4):
            raise ValueError(
                f"Parameter {row.param_name} has been input with "
                'value for "dist_param4" while '
                '"dist_param3" is empty. This is not '
                "allowed"
            )
        distparams = [
            item
            for item in [
                row.dist_param1,
                row.dist_param2,
                row.dist_param3,
                row.dist_param4,
            ]
            if _has_value(item)
        ]
        if "corr_sheet" in sensgroup:
            corrsheet = None if not _has_value(row.corr_sheet) else row.corr_sheet
        else:
            corrsheet = None
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams, corrsheet]

    return paramdict


def _raise_if_duplicates(container: Sequence[Hashable]) -> None:
    """Raises a descriptive error if there are duplicates in the container."""
    duplicates = {k: v for (k, v) in Counter(container).items() if v > 1}
    if duplicates:
        raise ValueError(f"Duplicates with counts: {duplicates}")


def _assert_no_merged_cells(input_filename: str) -> None:
    """Raises an exception if any merged cells exist, else returns None."""

    workbook = openpyxl.load_workbook(input_filename)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        merged_ranges = list(worksheet.merged_cells.ranges)
        if merged_ranges:
            raise ValueError(
                "Merged cells are not allowed. Found merged cell in "
                f"{input_filename} at sheet '{sheet_name}'.\n"
                f"Found {len(merged_ranges)} merged cell range(s): {merged_ranges}"
            )
